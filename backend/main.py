import os
import asyncio
import logging
from datetime import datetime, timedelta
from contextlib import asynccontextmanager

from fastapi import FastAPI, HTTPException, BackgroundTasks, Depends
from fastapi.middleware.cors import CORSMiddleware
from fastapi.staticfiles import StaticFiles
from fastapi.responses import Response, FileResponse
from sqlalchemy import select, desc, text

from database import init_db, AsyncSessionLocal, engine
from models import StockData, TurkishGold, ExchangeRate, AIInsight
from analysis import (
    analyze_stock, get_market_opportunities, get_bulk_analysis,
    BIST_SYMBOLS, GLOBAL_SYMBOLS, COMMODITIES_SYMBOLS
)
from ai_service import get_market_insight, get_stock_analysis
from workers import refresh_all_data, start_periodic_refresh

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s [%(name)s] %(levelname)s: %(message)s'
)
logger = logging.getLogger(__name__)

# ============================================================
# APP LIFECYCLE
# ============================================================

@asynccontextmanager
async def lifespan(app: FastAPI):
    """Startup and shutdown events"""
    logger.info("🐺 Wolfee Analytics starting up...")

    # Initialize database tables
    await init_db()
    logger.info("✅ Database initialized")

    # Start background refresh worker (10 min interval)
    refresh_task = asyncio.create_task(start_periodic_refresh(interval_minutes=10))
    logger.info("✅ Background refresh worker started (every 10 min)")

    yield

    # Shutdown
    refresh_task.cancel()
    logger.info("🛑 Wolfee Analytics shutting down")


app = FastAPI(title="Wolfee Analytics", lifespan=lifespan)

# ============================================================
# CORS
# ============================================================
origins = [
    "http://localhost:8080",
    "http://localhost:8000",
    "http://127.0.0.1:8080",
    "http://127.0.0.1:8000",
    "https://wolfee-backend.onrender.com",
    "https://solitary-scene-04bc.ozanggnr.workers.dev",
]

# Add Railway domain dynamically
railway_domain = os.getenv("RAILWAY_PUBLIC_DOMAIN")
if railway_domain:
    origins.append(f"https://{railway_domain}")

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],  # Allow all for flexibility
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)


# ============================================================
# HEALTH & ROOT
# ============================================================

@app.get("/")
@app.head("/")
async def read_root():
    """Serve frontend index.html or API welcome"""
    index_path = os.path.join(os.path.dirname(__file__), '..', 'frontend', 'index.html')
    if os.path.exists(index_path):
        return FileResponse(index_path)
    return {"message": "Welcome to Wolfee Analytics API"}


@app.get("/healthz")
@app.get("/health")
async def health_check():
    """Health check endpoint"""
    try:
        async with AsyncSessionLocal() as session:
            await session.execute(text("SELECT 1"))
        return {"status": "healthy", "service": "Wolfee Analytics API", "database": "connected"}
    except Exception:
        return {"status": "degraded", "service": "Wolfee Analytics API", "database": "disconnected"}


# ============================================================
# MARKET DATA ENDPOINTS (DB-first, instant response)
# ============================================================

@app.get("/api/stocks")
async def get_stocks():
    """Returns list of supported symbols."""
    return {
        "stocks": BIST_SYMBOLS + GLOBAL_SYMBOLS,
        "commodities": list(COMMODITIES_SYMBOLS.keys()),
        "bist_count": len(BIST_SYMBOLS),
        "global_count": len(GLOBAL_SYMBOLS),
        "commodity_count": len(COMMODITIES_SYMBOLS),
    }


@app.get("/api/market-data/quick")
async def get_quick_market_data(background_tasks: BackgroundTasks):
    """
    Fetch market data from PostgreSQL (instant).
    If DB is empty, triggers background refresh and returns what's available.
    """
    try:
        async with AsyncSessionLocal() as session:
            result = await session.execute(
                select(StockData).order_by(StockData.market_type, StockData.symbol)
            )
            stocks = result.scalars().all()

            if not stocks:
                # DB is empty — trigger refresh and return empty
                background_tasks.add_task(_sync_refresh)
                return {"stocks": [], "status": "loading", "message": "First load — data is being fetched. Refresh in 30 seconds."}

            stock_list = [_stock_to_dict(s) for s in stocks]
            return {"stocks": stock_list}

    except Exception as e:
        logger.error(f"Quick market data error: {e}")
        return {"stocks": [], "error": str(e)}


@app.get("/api/market-data/full")
async def get_full_market_data():
    """Returns ALL stocks from DB."""
    try:
        async with AsyncSessionLocal() as session:
            result = await session.execute(
                select(StockData).order_by(StockData.market_type, StockData.symbol)
            )
            stocks = result.scalars().all()
            return {"stocks": [_stock_to_dict(s) for s in stocks]}
    except Exception as e:
        logger.error(f"Full market data error: {e}")
        return {"stocks": []}


@app.get("/api/market-data")
async def get_market_data(background_tasks: BackgroundTasks):
    """Legacy endpoint."""
    return await get_quick_market_data(background_tasks)


# ============================================================
# TURKISH GOLD & EXCHANGE RATES
# ============================================================

@app.get("/api/turkish-gold")
async def get_turkish_gold():
    """Returns Turkish gold prices from DB."""
    try:
        async with AsyncSessionLocal() as session:
            result = await session.execute(select(TurkishGold))
            gold = result.scalars().all()
            return {
                "gold": [{
                    "gold_type": g.gold_type,
                    "display_name": g.display_name,
                    "buying_price": g.buying_price,
                    "selling_price": g.selling_price,
                    "change_pct": g.change_pct,
                    "updated_at": g.updated_at.isoformat() if g.updated_at else None
                } for g in gold]
            }
    except Exception as e:
        logger.error(f"Turkish gold error: {e}")
        return {"gold": []}


@app.get("/api/exchange-rates")
async def get_exchange_rates():
    """Returns exchange rates from DB."""
    try:
        async with AsyncSessionLocal() as session:
            result = await session.execute(select(ExchangeRate))
            rates = result.scalars().all()
            return {
                "rates": [{
                    "pair": r.pair,
                    "display_name": r.display_name,
                    "buying": r.buying,
                    "selling": r.selling,
                    "change_pct": r.change_pct,
                    "updated_at": r.updated_at.isoformat() if r.updated_at else None
                } for r in rates]
            }
    except Exception as e:
        logger.error(f"Exchange rates error: {e}")
        return {"rates": []}


# ============================================================
# STOCK ANALYSIS & HISTORY
# ============================================================

@app.get("/api/analyze/{symbol}")
async def get_stock_analysis_endpoint(symbol: str):
    """Returns analysis for a specific stock."""
    # Auto-append .IS if needed
    if "=" not in symbol and not symbol.endswith(".IS"):
        # Check if it's a global symbol
        if symbol.upper() not in [s.upper() for s in GLOBAL_SYMBOLS]:
            symbol += ".IS"

    is_commodity = "=" in symbol
    data = analyze_stock(symbol, is_commodity, detailed=True)
    if not data:
        raise HTTPException(status_code=404, detail="Stock data not found")
    return data


@app.get("/api/history/{symbol}")
async def get_stock_history(symbol: str, period: str = "1y"):
    """Returns historical price data for charts."""
    is_global = symbol.upper() in [s.upper() for s in GLOBAL_SYMBOLS]
    is_commodity = "=" in symbol

    if not is_global and not is_commodity and not symbol.endswith(".IS"):
        symbol += ".IS"

    try:
        if symbol.endswith(".IS"):
            from data_sources.turkish_market import fetch_bist_history
            data = fetch_bist_history(symbol, period)
        else:
            from data_sources.global_market import fetch_global_history
            data = fetch_global_history(symbol, period)

        if not data or not data.get('history'):
            raise HTTPException(status_code=404, detail="No history found")

        return {
            "symbol": symbol,
            "name": symbol,
            "history": data["history"]
        }
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"History error for {symbol}: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@app.get("/api/chart/{symbol}/{period}")
async def get_chart_data(symbol: str, period: str):
    """Chart data endpoint for frontend."""
    is_global = symbol.upper() in [s.upper() for s in GLOBAL_SYMBOLS]
    is_commodity = "=" in symbol

    if not symbol.endswith('.IS') and not is_global and "=" not in symbol:
        symbol += '.IS'

    try:
        if symbol.endswith(".IS"):
            from data_sources.turkish_market import fetch_bist_history
            data = fetch_bist_history(symbol, period)
        else:
            from data_sources.global_market import fetch_global_history
            data = fetch_global_history(symbol, period)

        if not data:
            raise HTTPException(status_code=404, detail="No chart data available")

        return {"history": data}

        return data

    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Chart error for {symbol}: {e}")
        raise HTTPException(status_code=404, detail="Chart data unavailable")


# ============================================================
# OPPORTUNITIES & AI INSIGHTS
# ============================================================

@app.get("/api/opportunities")
async def get_opportunities():
    """Returns buy opportunities from DB."""
    try:
        async with AsyncSessionLocal() as session:
            result = await session.execute(select(StockData))
            stocks = result.scalars().all()
            stock_list = [_stock_to_dict(s) for s in stocks]
            opps = get_market_opportunities(cached_data=stock_list)
            return {"opportunities": opps}
    except Exception as e:
        logger.error(f"Opportunities error: {e}")
        return {"opportunities": []}


@app.get("/api/insight")
async def get_insight():
    """Returns latest AI insight from DB."""
    try:
        async with AsyncSessionLocal() as session:
            result = await session.execute(
                select(AIInsight)
                .where(AIInsight.insight_type == 'daily')
                .order_by(desc(AIInsight.created_at))
                .limit(1)
            )
            insight = result.scalar_one_or_none()

            if insight and insight.insight_text:
                return {"insight": insight.insight_text}

            # Generate on-the-fly if no cached insight
            stock_result = await session.execute(select(StockData).limit(50))
            stocks = stock_result.scalars().all()
            stock_list = [_stock_to_dict(s) for s in stocks]
            insight_text = get_market_insight(stock_list)
            return {"insight": insight_text}

    except Exception as e:
        logger.error(f"Insight error: {e}")
        return {"insight": "🐺 Wolfee AI is warming up. Check back shortly."}


@app.get("/api/ai/analyze/{symbol}")
async def get_ai_stock_analysis(symbol: str):
    """Get Gemini AI deep analysis for a specific stock."""
    # Try to get stock data from DB first
    try:
        async with AsyncSessionLocal() as session:
            result = await session.execute(
                select(StockData).where(StockData.symbol == symbol)
            )
            stock = result.scalar_one_or_none()

            if stock:
                stock_data = _stock_to_dict(stock)
            else:
                # Fetch fresh
                stock_data = analyze_stock(symbol)

            if not stock_data:
                raise HTTPException(status_code=404, detail="Stock not found")

            import asyncio
            analysis = await asyncio.to_thread(get_stock_analysis, symbol, stock_data)
            return {"symbol": symbol, "analysis": analysis}

    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"AI analysis error for {symbol}: {e}")
        return {"symbol": symbol, "analysis": "AI analysis temporarily unavailable."}


# ============================================================
# MANUAL REFRESH
# ============================================================

@app.post("/api/refresh")
async def trigger_refresh(background_tasks: BackgroundTasks):
    """Manually trigger a full data refresh."""
    background_tasks.add_task(_sync_refresh)
    return {"status": "Refresh triggered", "message": "Data will be updated in the background."}


@app.get("/api/refresh")
async def trigger_refresh_get(background_tasks: BackgroundTasks):
    """GET version of refresh for easy browser access."""
    background_tasks.add_task(_sync_refresh)
    return {"status": "Refresh triggered"}


def _sync_refresh():
    """Run async refresh in sync context (for BackgroundTasks)."""
    loop = asyncio.new_event_loop()
    asyncio.set_event_loop(loop)
    try:
        loop.run_until_complete(refresh_all_data())
    finally:
        loop.close()


# ============================================================
# ROUTE ALIASES (backward compatibility)
# ============================================================

@app.get("/insight")
async def get_insight_alias():
    return await get_insight()


@app.get("/opportunities")
async def get_opportunities_alias():
    return await get_opportunities()


# ============================================================
# EXCEL EXPORT
# ============================================================

@app.get("/api/export/portfolio")
async def export_portfolio(symbols: str, period: str):
    """Export selected portfolio stocks to Excel."""
    try:
        if period not in ["daily", "weekly", "monthly"]:
            raise HTTPException(status_code=400, detail="Invalid period")

        symbol_list = [s.strip() for s in symbols.split(',')]
        if not symbol_list:
            raise HTTPException(status_code=400, detail="No symbols provided")

        results = []
        today = datetime.now().strftime("%Y-%m-%d")

        for symbol in symbol_list:
            data = analyze_stock(symbol, is_commodity="=" in symbol)
            if data:
                price = data.get('price', 0)
                change_p = data.get('change_pct', 0)
                change_amt = price * (change_p / 100) if price else 0
                start_price = data.get('previous_close') or (price - change_amt)

                rsi = data.get('rsi', 0)
                rsi_status = "Oversold" if rsi < 30 else "Overbought" if rsi > 70 else "Neutral"

                results.append({
                    "Symbol": data.get('symbol', ''),
                    "Name": data.get('name', ''),
                    "Report Period": "Session Snapshot",
                    "Analysis Date": today,
                    "Trend": data.get('prediction', ''),
                    "Current Price": round(price, 2),
                    "Start Price": round(start_price, 2),
                    "Change %": round(change_p, 2),
                    "Change Amt": round(change_amt, 2),
                    "Period High": data.get('day_high', 0),
                    "High Date": today,
                    "Period Low": data.get('day_low', 0),
                    "Low Date": today,
                    "RSI (14)": round(rsi, 2),
                    "RSI Status": rsi_status,
                    "Volatility %": data.get('volatility', 'MEDIUM'),
                    "MA(20)": data.get('ma_20', 0),
                    "Volume (Period)": data.get('volume', 0),
                })

        if not results:
            raise HTTPException(status_code=404, detail="No data found for provided symbols")

        return _create_excel_response(results, f"portfolio_{period}", f"Portfolio {period.capitalize()}")

    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Portfolio export error: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@app.get("/api/export/{period}")
async def export_analysis(period: str):
    """Generate professional Excel export."""
    if period not in ["daily", "weekly", "monthly"]:
        raise HTTPException(status_code=400, detail="Invalid period. Use daily, weekly, or monthly.")

    data = get_bulk_analysis(period)
    if not data:
        raise HTTPException(status_code=404, detail="No data available for export.")

    return _create_excel_response(data, f"wolfee_market_{period}", period.capitalize())

def _create_excel_response(data: list, filename: str, sheet_name: str) -> Response:
    """Create a professionally formatted Excel file."""
    import io
    import pandas as pd
    from openpyxl.styles import Alignment, Font, PatternFill, Border, Side, numbers

    df = pd.DataFrame(data)

    columns_order = [
        "Symbol", "Name", "Report Period", "Analysis Date", "Trend",
        "Current Price", "Start Price", "Change %", "Change Amt",
        "Period High", "High Date", "Period Low", "Low Date",
        "RSI (14)", "RSI Status", "Volatility %", "MA(20)", "Volume (Period)"
    ]

    for col in columns_order:
        if col not in df.columns:
            df[col] = ""

    df = df.reindex(columns=columns_order)

    output = io.BytesIO()

    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name=sheet_name)

        workbook = writer.book
        ws = writer.sheets[sheet_name]

        # Styles
        center = Alignment(horizontal='center', vertical='center', wrap_text=True)
        left = Alignment(horizontal='left', vertical='center')
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill(start_color="1B3A5C", end_color="1B3A5C", fill_type="solid")
        alt_fill = PatternFill(start_color="F0F4F8", end_color="F0F4F8", fill_type="solid")
        gain_font = Font(color="157A3B", bold=True)
        loss_font = Font(color="C0392B", bold=True)
        thin_border = Border(
            left=Side(style='thin', color='D0D5DD'),
            right=Side(style='thin', color='D0D5DD'),
            top=Side(style='thin', color='D0D5DD'),
            bottom=Side(style='thin', color='D0D5DD')
        )
        number_format_price = '#,##0.00'
        number_format_pct = '0.00"%"'
        number_format_int = '#,##0'

        # Format header row
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center
            cell.border = thin_border

        # Format data rows
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row), start=2):
            fill = alt_fill if row_idx % 2 == 0 else PatternFill()

            for col_idx, cell in enumerate(row, start=1):
                cell.alignment = center
                cell.border = thin_border
                if fill.fgColor and fill.fgColor.rgb != '00000000':
                    cell.fill = fill

                col_name = columns_order[col_idx - 1] if col_idx <= len(columns_order) else ""

                # Number formatting
                if col_name in ["Current Price", "Start Price", "Change Amt", "Period High", "Period Low", "MA(20)"]:
                    if isinstance(cell.value, (int, float)):
                        cell.number_format = number_format_price

                elif col_name == "Change %":
                    if isinstance(cell.value, (int, float)):
                        cell.number_format = '0.00'
                        if cell.value > 0:
                            cell.font = gain_font
                        elif cell.value < 0:
                            cell.font = loss_font

                elif col_name == "Volume (Period)":
                    if isinstance(cell.value, (int, float)):
                        cell.number_format = number_format_int

                elif col_name == "RSI (14)":
                    if isinstance(cell.value, (int, float)):
                        cell.number_format = '0.00'

                # Name column left-aligned
                if col_name in ["Name", "Symbol"]:
                    cell.alignment = left

                # Trend column coloring
                if col_name == "Trend":
                    val = str(cell.value or "")
                    if "UP" in val:
                        cell.font = gain_font
                    elif "DOWN" in val:
                        cell.font = loss_font

                # RSI Status coloring
                if col_name == "RSI Status":
                    val = str(cell.value or "")
                    if "Overbought" in val:
                        cell.font = loss_font
                    elif "Oversold" in val:
                        cell.font = gain_font

        # Auto-fit column widths
        for col_idx, column_cells in enumerate(ws.columns, start=1):
            max_length = 0
            for cell in column_cells:
                try:
                    cell_len = len(str(cell.value or ""))
                    if cell_len > max_length:
                        max_length = cell_len
                except:
                    pass
            from openpyxl.utils import get_column_letter
            col_letter = get_column_letter(col_idx)
            ws.column_dimensions[col_letter].width = min(max(max_length + 3, 10), 30)

        # Freeze top row
        ws.freeze_panes = 'A2'

        # Add auto-filter
        ws.auto_filter.ref = ws.dimensions

    output.seek(0)

    return Response(
        content=output.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f"attachment; filename={filename}.xlsx"
        }
    )


# ============================================================
# STATIC FILE SERVING (Frontend)
# ============================================================

# Mount static files LAST so API routes take priority
frontend_path = os.path.join(os.path.dirname(__file__), '..', 'frontend')
if os.path.exists(frontend_path):
    app.mount("/", StaticFiles(directory=frontend_path, html=True), name="frontend")


# ============================================================
# HELPERS
# ============================================================

def _stock_to_dict(stock: StockData) -> dict:
    """Convert StockData model to dict for API response."""
    return {
        "symbol": stock.symbol,
        "name": stock.name or stock.symbol,
        "price": stock.price or 0,
        "change_pct": stock.change_pct or 0,
        "volume": stock.volume or 0,
        "day_high": stock.day_high or 0,
        "day_low": stock.day_low or 0,
        "open": stock.open_price or 0,
        "previous_close": stock.previous_close or 0,
        "bid": stock.bid or 0,
        "ask": stock.ask or 0,
        "rsi": stock.rsi or 50,
        "ma_20": stock.ma_20 or 0,
        "volatility": stock.volatility or "LOW",
        "currency": stock.currency or "USD",
        "market_type": stock.market_type or "GLOBAL",
        "prediction": stock.prediction or "",
        "reason": stock.reason or "",
        "is_favorable": stock.is_favorable or False,
        "is_buyable": stock.is_buyable or False,
        "market_cap": stock.market_cap or 0,
    }
